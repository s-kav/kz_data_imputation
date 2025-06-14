# -*- coding: utf-8 -*-
import os 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

# books = os.listdir('../data/test') 
books = os.listdir('../data/full-292') 

count = 0
# gaps = []

left = 0
center = 0
right = 0

for book in books:
	# workbook = load_workbook('../data/test/' + book)
	workbook = load_workbook('../data/full-292/' + book)
	ws = workbook.active

	for i in range (5, 34):
		for j in range(3, 29):
			if ws.cell(i, j).value is None: break
	count += 1
print(count)



# вручную
# SE.SEC.ENRL.GC.FE.ZS.xlsx	

"""
#############################################
#				  Единичные					#
#############################################

if ws.cell(i, 3).value is None and ws.cell(i, 4).value is not None: count += 1
for j in range(4, 28):
	cl = ws.cell(i, j-1).value
	c1 = ws.cell(i, j).value
	cr = ws.cell(i, j+1).value
	if cl is not None and c1 is None and cr is not None: count += 1
if ws.cell(i, 28).value is None and ws.cell(i, 27).value is not None: count += 1

############################################
				  Двоичные					
############################################

c1 = ws.cell(i, 3).value
c2 = ws.cell(i, 4).value
cr = ws.cell(i, 5).value
if (c1 == c2 == None) and (cr is not None): count += 2

for j in range(4, 27):
	cl = ws.cell(i, j-1).value
	c1 = ws.cell(i, j).value
	c2 = ws.cell(i, j+1).value
	cr = ws.cell(i, j+2).value
	if (cl is not None) and (c1 == c2 == None) and (cr is not None): count += 2

c1 = ws.cell(i, 28).value
c2 = ws.cell(i, 27).value
cl = ws.cell(i, 26).value
if (cl is not None) and (c1 == c2 == None): count += 2


############################################
				  Троичные				
############################################

c1 = ws.cell(i, 3).value
c2 = ws.cell(i, 4).value
c3 = ws.cell(i, 5).value
cr = ws.cell(i, 6).value
if (c1 == c2 == c3 == None) and (cr is not None): count += 3

for j in range(5, 26):
	cl = ws.cell(i, j-2).value
	c1 = ws.cell(i, j-1).value
	c2 = ws.cell(i, j).value
	c3 = ws.cell(i, j+1).value
	cr = ws.cell(i, j+2).value
	if (cl != None) and (c1 == c2 == c3 == None) and (cr != None): count += 3

cl = ws.cell(i, 25).value
c1 = ws.cell(i, 26).value
c2 = ws.cell(i, 27).value
c3 = ws.cell(i, 28).value
if (cl is not None) and (c1 == c2 == c3 == None): count += 3

############################################
				  4				
############################################

c1 = ws.cell(i, 3).value
c2 = ws.cell(i, 4).value
c3 = ws.cell(i, 5).value
c4 = ws.cell(i, 6).value
cr = ws.cell(i, 7).value
if (c1 == c2 == c3 == c4 == None) and (cr is not None): count += 4

for j in range(4, 25):
	cl = ws.cell(i, j-1).value
	c1 = ws.cell(i, j).value
	c2 = ws.cell(i, j+1).value
	c3 = ws.cell(i, j+2).value
	c4 = ws.cell(i, j+3).value
	cr = ws.cell(i, j+4).value
	if (cl != None) and (c1 == c2 == c3 == c4 == None) and (cr != None): count += 4

cl = ws.cell(i, 24).value
c1 = ws.cell(i, 25).value
c2 = ws.cell(i, 26).value
c3 = ws.cell(i, 27).value
c4 = ws.cell(i, 28).value
if (cl is not None) and (c1 == c2 == c3 == c4 == None): count += 4

############################################
				  5		
############################################

c1 = ws.cell(i, 3).value
c2 = ws.cell(i, 4).value
c3 = ws.cell(i, 5).value
c4 = ws.cell(i, 6).value
c5 = ws.cell(i, 7).value
cr = ws.cell(i, 8).value
if (c1 == c2 == c3 == c4 == c5 == None) and (cr is not None): count += 5

for j in range(4, 24):
	cl = ws.cell(i, j-1).value
	c1 = ws.cell(i, j).value
	c2 = ws.cell(i, j+1).value
	c3 = ws.cell(i, j+2).value
	c4 = ws.cell(i, j+3).value
	c5 = ws.cell(i, j+4).value
	cr = ws.cell(i, j+5).value
	if (cl != None) and (c1 == c2 == c3 == c4 == c5 == None) and (cr != None): count += 5

cl = ws.cell(i, 23).value
c1 = ws.cell(i, 24).value
c2 = ws.cell(i, 25).value
c3 = ws.cell(i, 26).value
c4 = ws.cell(i, 27).value
c5 = ws.cell(i, 28).value
if (cl is not None) and (c1 == c2 == c3 == c4 == c5 == None): count += 5

# Поиск всех пропусков
for j in range(3, 29):
	if ws.cell(i, j).value is None: count += 1

"""
