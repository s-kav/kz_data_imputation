# -*- coding: utf-8 -*-
import os 
from openpyxl import load_workbook
 
books = os.listdir('../data/max-5-gaps-292') 

count = 0

for book in books:
	flag = False
	workbook = load_workbook('../data/max-5-gaps-292/' + book)
	ws = workbook.active

	for i in range (5, 34):
		for j in range(3, 29):
			if ws.cell(i, j).value is None: 
				flag = True

	if flag == False: count += 1

print(count)
