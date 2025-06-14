# -*- coding: utf-8 -*-
import os 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from statistics import mean

greyFill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

books = os.listdir('../data/after-2/') 

count = 0

for book in books:
	workbook = load_workbook('../data/after-2/' + book)
	# workbook = load_workbook("test.xlsx")
	ws = workbook.active

	for i in range (5, 34):
		begin = [ws.cell(i, j).value for j in range(4, 7)]
		if None not in begin and ws.cell(i, 3).value is None:
			ws.cell(i, 3).value = mean(begin)
			ws.cell(i, 3).fill = greyFill
			count += 1

		end = [ws.cell(i, j).value for j in range(25, 28)]
		if None not in end and ws.cell(i, 28).value is None:
			ws.cell(i, 28).value = mean(end)
			ws.cell(i, 28).fill = greyFill
			count += 1
		
		for k in range(4, 28):
			if ws.cell(i, k).value is None and ws.cell(i, k-1).value is not None and ws.cell(i, k+1).value is not None:
				ws.cell(i, k).value = (ws.cell(i, k-1).value + ws.cell(i, k+1).value) / 2
				ws.cell(i, k).fill = greyFill
				count += 1

	workbook.save('../data/after-2/' + book)
print(count)
# workbook.save("test.xlsx")
