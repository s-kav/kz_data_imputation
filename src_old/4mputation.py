# -*- coding: utf-8 -*-
import os 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from statistics import mean


yellowFill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')

books = os.listdir('../data/after-4/') 

count = 0

for book in books:
	workbook = load_workbook('../data/after-4/' + book)
	# workbook = load_workbook("test.xlsx")
	ws = workbook.active

	for i in range (5, 34):

		begin = [ws.cell(i, j).value for j in range(3, 7)]
		if all(k is None for k in begin):
			arr = [ws.cell(i, j).value for j in range(7, 12)]
			if None not in arr: 
				m = mean(arr)
				for j in range(3, 7):
					ws.cell(i, j).value = m
					ws.cell(i, j).fill = yellowFill
				count += 4
			
		for k in range(4, 25):
			center = [ws.cell(i, j).value for j in range(k, k+4)]

			if all(c is None for c in center):
				arrLeft  = [ws.cell(i, j).value for j in range(3, k)]
				arrRight = [ws.cell(i, j).value for j in range(k+4, 29)]
				if k <= 8: arrRight = arrRight[:5]
				elif k >= 21: arrLeft  = arrLeft[-5:]
				else:
					arrLeft  = arrLeft[-5:]
					arrRight = arrRight[:5]

				if None not in arrLeft and None not in arrRight:
					ws.cell(i, k).value   = mean(arrLeft)
					ws.cell(i, k+3).value = mean(arrRight)

					l3 = ws.cell(i, k-3).value
					l2 = ws.cell(i, k-2).value
					l1 = ws.cell(i, k-1).value 
					a = ws.cell(i, k).value
					b = ws.cell(i, k+1).value
					c = ws.cell(i, k+2).value
					d = ws.cell(i, k+3).value
					r1 = ws.cell(i, k+4).value
					r2 = ws.cell(i, k+5).value
					r3 = ws.cell(i, k+6).value

					if 12 <= k+3 <= 19:
						ws.cell(i, k+1).value = (l2 + l1 + a + d + r1) / 5
						ws.cell(i, k+2).value = (l1 + a + d + r1 + r2) / 5
					else:
						ws.cell(i, k+1).value = (l3 + l2 + l3 + a + d) / 5
						ws.cell(i, k+2).value = (a + d + r1 + r2 + r3) / 5

					for j in range(k, k+4):
						ws.cell(i, j).fill = yellowFill
					count += 4

	workbook.save('../data/after-4/' + book)

print(count)









	# Хужий способ для средних

	# for k in range(4, 25):
	# 	center = [ws.cell(i, j).value for j in range(k, k+3)]
	# 	if all(c is None for c in center):
	# 		if 7 < k < 22:
	# 			arrLeft = [ws.cell(i, j).value for j in range(k-5, k)]
	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, k+8)]
	# 		elif k <= 7:
	# 			arrLeft = [ws.cell(i, j).value for j in range(3, k)]
	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, 29)]
	# 			print(arrLeft)
	# 			print(arrRight)
				

	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, 29)]
				