# -*- coding: utf-8 -*-
import os 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from statistics import mean

blueFill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

books = os.listdir('../data/after-2/') 

count = 0

def fill(i, k, leftArr, rightArr):
	global count
	ws.cell(i, k).value = mean(leftArr)
	ws.cell(i, k+1).value = mean(rightArr)
	ws.cell(i, k).fill = blueFill
	ws.cell(i, k+1).fill = blueFill
	count += 2

for book in books:
	workbook = load_workbook('../data/after-2/' + book)
	# workbook = load_workbook("test.xlsx")
	ws = workbook.active


	for i in range (5, 34):

		if ws.cell(i, 3).value is None and ws.cell(i, 4).value is None:
			arr = [ws.cell(i, j).value for j in range(5, 10)]
			if None not in arr:
				fill(i, 3, arr, arr)

		if ws.cell(i, 27).value is None and ws.cell(i, 28).value is None:
			arr = [ws.cell(i, j).value for j in range(22, 27)]
			if None not in arr: 
				fill(i, 27, arr, arr)
		
		for k in range(4, 27):
			if  ws.cell(i, k).value is None and ws.cell(i, k+1).value is None:

				leftArr = []
				rightArr = []

				for j in range(k-1, 2, -1):
					if ws.cell(i, j).value is None: break 
					else: leftArr.append(ws.cell(i, j).value)
				for j in range(k+2, 29):
					if ws.cell(i, j).value is None: break 
					else: rightArr.append(ws.cell(i, j).value)

				if len(leftArr) > len(rightArr) and len(leftArr) > 2 and len(rightArr) > 1:
					leftArr = leftArr[:3] 
					rightArr = rightArr[:2]
					fill(i, k, leftArr, rightArr)

				elif len(leftArr) < len(rightArr) and len(leftArr) > 1 and len(rightArr) > 2:
					leftArr = leftArr[:2]
					rightArr = rightArr[:3]
					fill(i, k, leftArr, rightArr)

				elif len(leftArr) == len(rightArr) and len(leftArr) > 2 and len(rightArr) > 2:
					leftArr = leftArr[:3]
					rightArr = rightArr[:3]
					fill(i, k, leftArr, rightArr)

				# Два случая с 1 слева и справа
				if k == 4:
					arr = [ws.cell(i, j).value for j in range(6, 9)]
					if None not in arr and ws.cell(i, 3).value is not None: 
						fill(i, k, [ws.cell(i, 3).value], arr)
				if k == 26:
					arr = [ws.cell(i, j).value for j in range(25, 22, -1)]
					if None not in arr and ws.cell(i, 28).value is not None: 
						fill(i, k, arr, [ws.cell(i, 28).value])

	workbook.save('../data/after-2/' + book)


print(count)
# workbook.save('./ws2xlsx after imputation/' + book)

# Оценивать крыло до пропуска?
