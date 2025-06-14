# -*- coding: utf-8 -*-
import os 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from statistics import mean


greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

books = os.listdir('../data/after-3/') 

count = 0

for book in books:
	workbook = load_workbook('../data/after-3/' + book)
	# workbook = load_workbook("test.xlsx")
	ws = workbook.active

	for i in range (5, 34):

		begin = [ws.cell(i, j).value for j in range(3, 6)]
		if all(k is None for k in begin):
			arr = [ws.cell(i, j).value for j in range(6, 11)]
			if None not in arr: 
				m = mean(arr)
				for j in range(3, 6):
					ws.cell(i, j).value = m
					ws.cell(i, j).fill = greenFill
				count += 3
			
		end =  [ws.cell(i, j).value for j in range(26, 29)]
		if all(k is None for k in end):
			arr = [ws.cell(i, j).value for j in range(21, 26)]
			if None not in arr: 
				m = mean(arr)
				for j in range(26, 29):
					ws.cell(i, j).value = m
					ws.cell(i, j).fill = greenFill
				count += 3
			
		for k in range(4, 26):
			center = [ws.cell(i, j).value for j in range(k, k+3)]
			if all(c is None for c in center):
				arrLeft  = [ws.cell(i, j).value for j in range(3, k)]
				arrRight = [ws.cell(i, j).value for j in range(k+3, 29)]
				if k <= 7: arrRight = arrRight[:5]
				elif k >= 22: arrLeft  = arrLeft[-5:]
				else:
					arrLeft  = arrLeft[-5:]
					arrRight = arrRight[:5]

				if None not in arrLeft and None not in arrRight:
					ws.cell(i, k).value   = mean(arrLeft)
					ws.cell(i, k+2).value = mean(arrRight)
					ws.cell(i, k+1).value = (ws.cell(i, k).value + ws.cell(i, k+2).value) / 2
					for j in range(k, k+3):
						ws.cell(i, j).fill = greenFill
					count += 3

	workbook.save('../data/after-3/' + book)

print(count)









	# Хужий способ для средних

	# for k in range(4, 25):
	# 	center = [ws.cell(i, j).value for j in range(k, k+3)]
	# 	if all(c is None for c in center):
	# 		if 7 < k < 22:
	# 			arrLeft = [ws.cell(i, j).value for j in range(k-5, k)]
	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, k+8)]
	# 		elif k <= 7:
	# 			arrLeft = [ws.cell(i, j).value for j in range(3, k)]
	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, 29)]
	# 			print(arrLeft)
	# 			print(arrRight)
				

	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, 29)]
				