# -*- coding: utf-8 -*-
import os 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from statistics import mean


redFill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')

books = os.listdir('../data/after-5/') 

count = 0

for book in books:
	workbook = load_workbook('../data/after-5/' + book)
	# workbook = load_workbook("test.xlsx")
	ws = workbook.active

	for i in range (5, 34):

		end =  [ws.cell(i, j).value for j in range(24, 29)]
		if all(k is None for k in end):
			arr = [ws.cell(i, j).value for j in range(19, 24)]
			if None not in arr: 
				m = mean(arr)
				for j in range(24, 29):
					ws.cell(i, j).value = m
					ws.cell(i, j).fill = redFill
				count += 5

	workbook.save('../data/after-5/' + book)

print(count)









	# Хужий способ для средних

	# for k in range(4, 25):
	# 	center = [ws.cell(i, j).value for j in range(k, k+3)]
	# 	if all(c is None for c in center):
	# 		if 7 < k < 22:
	# 			arrLeft = [ws.cell(i, j).value for j in range(k-5, k)]
	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, k+8)]
	# 		elif k <= 7:
	# 			arrLeft = [ws.cell(i, j).value for j in range(3, k)]
	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, 29)]
	# 			print(arrLeft)
	# 			print(arrRight)
				

	# 			arrRight = [ws.cell(i, j).value for j in range(k+3, 29)]
				