# -*- coding: utf-8 -*-

'''
Очищает полную выборку (1511 файлов) до тех, 
которые имеют не более 5 пропусков в строке
'''

import os 
import threading
from openpyxl import load_workbook

books = os.listdir('D:/Google Диск/институт/диплом/data2xlsx after 5 empty filter') 

count = 0
gaps=[]

for book in books:
	workbook = load_workbook('data2xlsx after 5 empty filter/' + book)
	data = workbook.active

	for i in range (5, 34):
		gaps = 0
		for j in range(3, 29):
			if data.cell(i, j).value is None:
				gaps += 1
		if gaps > 5:
			os.remove('data2xlsx after 5 empty filter/' + book)
			break


	# workbook.save('./data2xlsx after imputation/' + book)